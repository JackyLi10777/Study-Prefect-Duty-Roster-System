"""Bounded local CSV/XLSX parsing for operator-reviewed prefect imports."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
import re
from typing import Any
from zipfile import BadZipFile, LargeZipFile, ZipFile

from openpyxl import load_workbook

from nicegui_app.utils.prefect_import import FIELD_ALIASES
from roster_policy import is_chinese_display_name


MAX_IMPORT_BYTES = 2 * 1024 * 1024
MAX_IMPORT_ROWS = 2_000
MAX_IMPORT_COLUMNS = 50
MAX_XLSX_ARCHIVE_MEMBERS = 512
MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES = 32 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 250.0
MIN_XLSX_RATIO_CHECK_BYTES = 128 * 1024
REQUIRED_XLSX_MEMBERS = frozenset({"[Content_Types].xml", "xl/workbook.xml"})
TARGET_FIELDS = ("name_zh", "form", "class_name", "role", "available_days", "remarks")
REQUIRED_TARGET_FIELDS = frozenset({"name_zh", "form", "class_name", "role", "available_days"})

_HEADER_LABEL_MARKERS = (
    "姓名",
    "中文",
    "學生",
    "級別",
    "年級",
    "班別",
    "班級",
    "職務",
    "職位",
    "角色",
    "身份",
    "可值班",
    "可用",
    "備註",
    "導學風紀",
)


class PrefectFileImportError(ValueError):
    """Raised when an uploaded file cannot be safely previewed."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        super().__init__(message)


@dataclass(frozen=True)
class ParsedImportFile:
    filename: str
    sheet_name: str | None
    sheet_names: tuple[str, ...]
    headers: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]


@dataclass(frozen=True)
class ColumnProfile:
    header: str
    value_kinds: tuple[str, ...]
    non_empty_count: int


def parse_prefect_file(
    filename: str,
    content: bytes,
    *,
    sheet_name: str | None = None,
) -> ParsedImportFile:
    if not content:
        raise PrefectFileImportError("empty_file", "The selected file is empty.")
    if len(content) > MAX_IMPORT_BYTES:
        raise PrefectFileImportError("too_large", "The selected file is larger than 2 MB.")
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        return _parse_csv(filename, content)
    if suffix == ".xlsx":
        return _parse_xlsx(filename, content, sheet_name=sheet_name)
    if suffix in {".xls", ".xlsm"}:
        raise PrefectFileImportError(
            "plain_xlsx_required",
            "Save legacy or macro-enabled workbooks as a plain .xlsx file first.",
        )
    raise PrefectFileImportError("unsupported_format", "Only .csv and .xlsx prefect files are accepted.")


def suggest_local_column_mapping(headers: tuple[str, ...]) -> dict[str, str]:
    normalized_headers = {_normalize_header(header): header for header in headers}
    mapping: dict[str, str] = {}
    for target in TARGET_FIELDS:
        for alias in FIELD_ALIASES[target]:
            source = normalized_headers.get(_normalize_header(alias))
            if source is not None:
                mapping[target] = source
                break
    return mapping


def profile_columns(parsed: ParsedImportFile) -> tuple[ColumnProfile, ...]:
    profiles: list[ColumnProfile] = []
    for header in parsed.headers:
        non_empty = [row.get(header) for row in parsed.rows if row.get(header) not in (None, "")]
        kinds = sorted({_value_kind(value) for value in non_empty})
        profiles.append(ColumnProfile(header, tuple(kinds or ["EMPTY"]), len(non_empty)))
    return tuple(profiles)


def validate_target_mapping(target_to_source: dict[str, str], headers: tuple[str, ...]) -> dict[str, str]:
    unknown_targets = set(target_to_source) - set(TARGET_FIELDS)
    if unknown_targets:
        raise PrefectFileImportError("mapping_target", "The column mapping contains an unsupported target field.")
    missing = REQUIRED_TARGET_FIELDS - set(target_to_source)
    if missing:
        raise PrefectFileImportError(
            "mapping_required",
            "Map Chinese name, form, class, role, and available days before previewing.",
        )
    sources = list(target_to_source.values())
    if len(sources) != len(set(sources)):
        raise PrefectFileImportError("mapping_duplicate", "Each source column can be mapped only once.")
    if any(source not in headers for source in sources):
        raise PrefectFileImportError(
            "mapping_source",
            "The column mapping refers to a source column that is not in this file.",
        )
    return dict(target_to_source)


def _parse_csv(filename: str, content: bytes) -> ParsedImportFile:
    text = _decode_csv(content)
    sample = text[:8_192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(StringIO(text), dialect=dialect)
    try:
        raw_headers = next(reader)
    except StopIteration as error:
        raise PrefectFileImportError("empty_file", "The selected file is empty.") from error
    _reject_probable_data_row(raw_headers)
    headers = _validate_headers(tuple(str(header or "").strip() for header in raw_headers))
    rows: list[dict[str, Any]] = []
    for index, raw_values in enumerate(reader, start=1):
        if index > MAX_IMPORT_ROWS:
            raise PrefectFileImportError("too_many_rows", "The file contains more than 2,000 data rows.")
        if len(raw_values) > len(headers) and any(str(value).strip() for value in raw_values[len(headers) :]):
            raise PrefectFileImportError(
                "extra_csv_values",
                "A CSV row contains values beyond the declared column headings.",
            )
        values = [_clean_cell(value) for value in raw_values[: len(headers)]]
        values.extend([None] * (len(headers) - len(values)))
        row = dict(zip(headers, values, strict=True))
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    if not rows:
        raise PrefectFileImportError("no_data_rows", "The file contains no data rows.")
    return ParsedImportFile(filename, None, (), headers, tuple(rows))


def _parse_xlsx(filename: str, content: bytes, *, sheet_name: str | None) -> ParsedImportFile:
    _inspect_xlsx_archive(content)
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False, keep_links=False)
    except Exception as error:
        raise PrefectFileImportError("workbook_unreadable", "The .xlsx workbook could not be opened.") from error
    try:
        sheet_names = tuple(workbook.sheetnames)
        if not sheet_names:
            raise PrefectFileImportError("no_worksheets", "The workbook has no worksheets.")
        selected_name = sheet_name or sheet_names[0]
        if selected_name not in sheet_names:
            raise PrefectFileImportError("worksheet_missing", "The selected worksheet no longer exists.")
        worksheet = workbook[selected_name]
        iterator = worksheet.iter_rows()
        try:
            header_cells = next(iterator)
        except StopIteration as error:
            raise PrefectFileImportError("worksheet_empty", "The selected worksheet is empty.") from error
        if any(cell.data_type == "f" for cell in header_cells):
            raise PrefectFileImportError("formula_cells", "Formula cells are not accepted in prefect imports.")
        _reject_probable_data_row([cell.value for cell in header_cells])
        headers = _validate_headers(tuple(str(cell.value or "").strip() for cell in header_cells))
        rows: list[dict[str, Any]] = []
        for index, cells in enumerate(iterator, start=1):
            if index > MAX_IMPORT_ROWS:
                raise PrefectFileImportError("too_many_rows", "The worksheet contains more than 2,000 data rows.")
            if any(cell.data_type == "f" for cell in cells):
                raise PrefectFileImportError("formula_cells", "Formula cells are not accepted in prefect imports.")
            values = [_clean_cell(cell.value) for cell in cells[: len(headers)]]
            values.extend([None] * (len(headers) - len(values)))
            row = dict(zip(headers, values, strict=True))
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        if not rows:
            raise PrefectFileImportError("no_data_rows", "The selected worksheet contains no data rows.")
        return ParsedImportFile(filename, selected_name, sheet_names, headers, tuple(rows))
    finally:
        workbook.close()


def _inspect_xlsx_archive(content: bytes) -> None:
    """Reject malformed or expansion-heavy OOXML containers before openpyxl reads them."""

    try:
        with ZipFile(BytesIO(content), "r") as archive:
            members = archive.infolist()
            if not members or len(members) > MAX_XLSX_ARCHIVE_MEMBERS:
                raise _unsafe_workbook_archive()

            member_names = [member.filename for member in members]
            if len(member_names) != len(set(member_names)):
                raise _unsafe_workbook_archive()
            if not REQUIRED_XLSX_MEMBERS.issubset(member_names):
                raise _unsafe_workbook_archive()

            total_uncompressed = 0
            for member in members:
                if member.flag_bits & 0x1:
                    raise _unsafe_workbook_archive()
                if member.file_size < 0 or member.compress_size < 0:
                    raise _unsafe_workbook_archive()
                if member.file_size > MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES:
                    raise _unsafe_workbook_archive()
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES:
                    raise _unsafe_workbook_archive()
                if member.file_size >= MIN_XLSX_RATIO_CHECK_BYTES:
                    compression_ratio = member.file_size / max(member.compress_size, 1)
                    if compression_ratio > MAX_XLSX_COMPRESSION_RATIO:
                        raise _unsafe_workbook_archive()
    except PrefectFileImportError:
        raise
    except (BadZipFile, LargeZipFile, OSError, RuntimeError, ValueError) as error:
        raise _unsafe_workbook_archive() from error


def _unsafe_workbook_archive() -> PrefectFileImportError:
    return PrefectFileImportError(
        "unsafe_workbook_archive",
        "The .xlsx archive is invalid, encrypted, or exceeds safe expansion limits.",
    )


def _reject_probable_data_row(values: list[Any] | tuple[Any, ...]) -> None:
    """Keep a headerless first student row from becoming headings or API metadata."""

    texts = tuple(str(_clean_cell(value) or "").strip() for value in values)
    has_name = any(_looks_like_chinese_name_value(text) for text in texts)
    has_academic_value = any(_looks_like_form_value(text) or _looks_like_class_value(text) for text in texts)
    has_duty_value = any(_looks_like_role_value(text) or _looks_like_day_value(text) for text in texts)
    if has_name and has_academic_value and has_duty_value:
        raise PrefectFileImportError(
            "headings_required",
            "The first row appears to contain prefect data; add column headings before importing.",
        )


def _looks_like_chinese_name_value(text: str) -> bool:
    if not text or any(marker in text for marker in _HEADER_LABEL_MARKERS):
        return False
    return is_chinese_display_name(text)


def _looks_like_form_value(text: str) -> bool:
    normalized = re.sub(r"\s+", "", text.upper())
    return re.fullmatch(r"(?:FORM)?F?\.?[3-6]", normalized) is not None or normalized in {
        "中三",
        "中四",
        "中五",
        "中六",
    }


def _looks_like_class_value(text: str) -> bool:
    normalized = re.sub(r"[\s.-]+", "", text.upper())
    return re.fullmatch(r"(?:F?[3-6])(?:[A-Z]|[甲乙丙丁戊己庚辛])(?:班)?", normalized) is not None


def _looks_like_role_value(text: str) -> bool:
    normalized = text.casefold()
    return "study prefect" in normalized or "assistant head" in normalized or "導學風紀" in text


def _looks_like_day_value(text: str) -> bool:
    normalized = text.upper()
    return bool(
        re.search(r"(?:星期|週)[一二三四五]", text)
        or re.search(r"\b(?:MON(?:DAY)?|TUE(?:SDAY)?|WED(?:NESDAY)?|THU(?:RSDAY)?|FRI(?:DAY)?)\b", normalized)
    )


def _validate_headers(headers: tuple[str, ...]) -> tuple[str, ...]:
    if not headers or not any(headers):
        raise PrefectFileImportError("headings_required", "The first row must contain column headings.")
    if len(headers) > MAX_IMPORT_COLUMNS:
        raise PrefectFileImportError("too_many_columns", "The file contains more than 50 columns.")
    if any(not header for header in headers):
        raise PrefectFileImportError("blank_heading", "Every imported column must have a heading.")
    if len(headers) != len(set(headers)):
        raise PrefectFileImportError("duplicate_headings", "Column headings must be unique.")
    return headers


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp950"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise PrefectFileImportError(
        "unsupported_encoding",
        "The CSV encoding is not supported; save it as UTF-8 CSV.",
    )


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _normalize_header(value: str) -> str:
    return re.sub(r"[^0-9a-z\u3400-\u9fff]", "", value.casefold())


def _value_kind(value: Any) -> str:
    if isinstance(value, bool):
        return "BOOLEAN"
    if isinstance(value, (int, float)):
        return "NUMBER"
    text = str(value).strip()
    upper = text.upper()
    if re.fullmatch(r"(?:FORM\s*)?F?\.?[3-6]", upper):
        return "FORM_CODE"
    if any(token in upper for token in ("MON", "TUE", "WED", "THU", "FRI", "星期", "週")):
        return "DAY_LIST"
    if re.search(r"[\u3400-\u9fff]", text):
        return "CJK_TEXT"
    return "ASCII_TEXT"


__all__ = [
    "ColumnProfile",
    "MAX_IMPORT_BYTES",
    "MAX_IMPORT_COLUMNS",
    "MAX_IMPORT_ROWS",
    "ParsedImportFile",
    "PrefectFileImportError",
    "REQUIRED_TARGET_FIELDS",
    "TARGET_FIELDS",
    "parse_prefect_file",
    "profile_columns",
    "suggest_local_column_mapping",
    "validate_target_mapping",
]
