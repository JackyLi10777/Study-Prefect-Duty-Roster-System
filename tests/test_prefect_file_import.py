from __future__ import annotations

from io import BytesIO
import json
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
import pytest

import nicegui_app.utils.prefect_file_import as prefect_file_import
from nicegui_app.services.prefect_import_assistant import (
    ImportAssistantError,
    import_assistant_status,
    suggest_deepseek_column_mapping,
)
from nicegui_app.utils.prefect_file_import import (
    PrefectFileImportError,
    parse_prefect_file,
    profile_columns,
    suggest_local_column_mapping,
    validate_target_mapping,
)
from nicegui_app.utils.prefect_import import parse_prefect_import_rows


CSV_CONTENT = (
    "姓名,級別,班別,職務,可值班日,備註\n"
    "測試甲,F.3,3A,導學風紀,星期一、星期三,\n"
    "測試乙,F.5,5B,助理首席導學風紀,星期二、星期四,支援組長\n"
).encode("utf-8")


def _workbook_bytes(*, formula: bool = False) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "名冊"
    worksheet.append(["中文姓名", "Form", "Class", "Role", "Available days"])
    worksheet.append(["測試丙", "F.4", "4C", "Study Prefect", "Monday Wednesday"])
    if formula:
        worksheet["B2"] = "=1+2"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _headerless_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["陳大文", "F.4", "4C", "Study Prefect", "Monday Wednesday"])
    worksheet.append(["李小明", "F.5", "5A", "Study Prefect", "Tuesday Thursday"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _archive_bytes(entries: list[tuple[str, bytes]]) -> bytes:
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        for filename, content in entries:
            archive.writestr(filename, content)
    return output.getvalue()


def _encrypted_archive_bytes() -> bytes:
    content = bytearray(
        _archive_bytes(
            [
                ("[Content_Types].xml", b"<Types/>"),
                ("xl/workbook.xml", b"<workbook/>"),
            ]
        )
    )
    local_header = content.find(b"PK\x03\x04")
    central_header = content.find(b"PK\x01\x02")
    assert local_header >= 0 and central_header >= 0
    local_flags = int.from_bytes(content[local_header + 6 : local_header + 8], "little") | 0x1
    central_flags = int.from_bytes(content[central_header + 8 : central_header + 10], "little") | 0x1
    content[local_header + 6 : local_header + 8] = local_flags.to_bytes(2, "little")
    content[central_header + 8 : central_header + 10] = central_flags.to_bytes(2, "little")
    return bytes(content)


def test_csv_file_is_parsed_locally_and_reuses_existing_prefect_validation() -> None:
    parsed = parse_prefect_file("prefects.csv", CSV_CONTENT)
    mapping = suggest_local_column_mapping(parsed.headers)
    validate_target_mapping(mapping, parsed.headers)

    preview = parse_prefect_import_rows(list(parsed.rows), target_to_source=mapping)

    assert parsed.filename == "prefects.csv"
    assert len(parsed.rows) == 2
    assert preview.issues == ()
    assert [row.name_zh for row in preview.rows] == ["測試甲", "測試乙"]
    assert [row.role_code for row in preview.rows] == ["study_prefect", "assistant_head"]


def test_csv_header_whitespace_is_removed_without_losing_column_values() -> None:
    parsed = parse_prefect_file(
        "prefects.csv",
        " 姓名 , 級別 , 班別 , 職務 , 可值班日 \n測試風紀,F.3,3H,導學風紀,星期一".encode("utf-8"),
    )
    mapping = suggest_local_column_mapping(parsed.headers)
    preview = parse_prefect_import_rows(list(parsed.rows), target_to_source=mapping)

    assert parsed.headers == ("姓名", "級別", "班別", "職務", "可值班日")
    assert parsed.rows[0]["姓名"] == "測試風紀"
    assert preview.issues == ()
    assert preview.rows[0].name_zh == "測試風紀"


@pytest.mark.parametrize(
    ("content", "expected_code"),
    [
        ("姓名,,班別,職務,可值班日\n測試甲,F.3,3A,導學風紀,星期一", "blank_heading"),
        ("姓名,姓名,班別,職務,可值班日\n測試甲,F.3,3A,導學風紀,星期一", "duplicate_headings"),
    ],
)
def test_file_csv_rejects_blank_and_duplicate_headings(content: str, expected_code: str) -> None:
    with pytest.raises(PrefectFileImportError) as captured:
        parse_prefect_file("prefects.csv", content.encode("utf-8"))

    assert captured.value.code == expected_code


def test_xlsx_uses_a_named_sheet_without_executing_formulas() -> None:
    parsed = parse_prefect_file("prefects.xlsx", _workbook_bytes(), sheet_name="名冊")
    mapping = suggest_local_column_mapping(parsed.headers)
    preview = parse_prefect_import_rows(list(parsed.rows), target_to_source=mapping)

    assert parsed.sheet_name == "名冊"
    assert parsed.sheet_names == ("名冊",)
    assert preview.issues == ()
    assert preview.rows[0].name_zh == "測試丙"

    with pytest.raises(PrefectFileImportError, match="Formula"):
        parse_prefect_file("formula.xlsx", _workbook_bytes(formula=True))


@pytest.mark.parametrize(
    ("filename", "content"),
    [
        ("headerless.csv", "陳大文,F.4,4C,導學風紀,星期一、星期三\n李小明,F.5,5A,導學風紀,星期二".encode("utf-8")),
        ("headerless.xlsx", _headerless_workbook_bytes()),
    ],
)
def test_headerless_roster_is_rejected_before_any_external_mapping_request(
    filename: str,
    content: bytes,
    monkeypatch,
) -> None:
    monkeypatch.setenv("SING_YIN_DEEPSEEK_ENABLED", "true")
    monkeypatch.setenv("SING_YIN_DEEPSEEK_API_KEY", "fake-local-test-key")
    transport_calls: list[dict[str, object]] = []

    def transport(payload, _api_key, _timeout):
        transport_calls.append(payload)
        return {"choices": [{"message": {"content": '{"mapping":[]}'}}]}

    with pytest.raises(PrefectFileImportError) as captured:
        parsed = parse_prefect_file(filename, content)
        suggest_deepseek_column_mapping(parsed, transport=transport)

    assert captured.value.code == "headings_required"
    assert transport_calls == []


@pytest.mark.parametrize(
    ("limit_name", "limit_value"),
    [
        ("MAX_XLSX_ARCHIVE_MEMBERS", 1),
        ("MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES", 1),
        ("MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES", 1),
    ],
)
def test_xlsx_archive_metadata_caps_run_before_openpyxl(
    limit_name: str,
    limit_value: int,
    monkeypatch,
) -> None:
    monkeypatch.setattr(prefect_file_import, limit_name, limit_value)
    monkeypatch.setattr(
        prefect_file_import,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("openpyxl must not read a rejected archive"),
    )

    with pytest.raises(PrefectFileImportError) as captured:
        parse_prefect_file("bounded.xlsx", _workbook_bytes())

    assert captured.value.code == "unsafe_workbook_archive"


def test_high_compression_ratio_xlsx_is_rejected_before_openpyxl(monkeypatch) -> None:
    compressed_bomb = _archive_bytes(
        [
            ("[Content_Types].xml", b"<Types/>"),
            ("xl/workbook.xml", b"<workbook/>"),
            ("xl/worksheets/sheet1.xml", b"A" * (512 * 1024)),
        ]
    )
    assert len(compressed_bomb) < prefect_file_import.MAX_IMPORT_BYTES
    monkeypatch.setattr(
        prefect_file_import,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("openpyxl must not read a suspicious archive"),
    )

    with pytest.raises(PrefectFileImportError) as captured:
        parse_prefect_file("compressed.xlsx", compressed_bomb)

    assert captured.value.code == "unsafe_workbook_archive"


@pytest.mark.parametrize("content", [b"not-a-zip", _encrypted_archive_bytes()])
def test_invalid_or_encrypted_xlsx_uses_one_stable_error_code(content: bytes, monkeypatch) -> None:
    monkeypatch.setattr(
        prefect_file_import,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("openpyxl must not read an unsafe archive"),
    )

    with pytest.raises(PrefectFileImportError) as captured:
        parse_prefect_file("unsafe.xlsx", content)

    assert captured.value.code == "unsafe_workbook_archive"


def test_unsafe_or_unsupported_workbooks_fail_with_a_clear_next_step() -> None:
    with pytest.raises(PrefectFileImportError, match="plain .xlsx"):
        parse_prefect_file("legacy.xlsm", b"not-a-workbook")
    with pytest.raises(PrefectFileImportError, match="Only .csv and .xlsx"):
        parse_prefect_file("prefects.json", b"[]")
    with pytest.raises(PrefectFileImportError, match="larger than 2 MB"):
        parse_prefect_file("large.csv", b"x" * (2 * 1024 * 1024 + 1))


def test_csv_and_xlsx_share_the_same_cell_length_boundary() -> None:
    oversized_cell = "字" * (prefect_file_import.MAX_IMPORT_CELL_CHARACTERS + 1)
    csv_content = (
        "姓名,級別,班別,職務,可值班日,備註\n"
        f"測試甲,F.3,3A,導學風紀,星期一,{oversized_cell}\n"
    ).encode("utf-8")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["姓名", "級別", "班別", "職務", "可值班日", "備註"])
    worksheet.append(["測試甲", "F.3", "3A", "導學風紀", "星期一", oversized_cell])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    for filename, content in (("long.csv", csv_content), ("long.xlsx", output.getvalue())):
        with pytest.raises(PrefectFileImportError) as captured:
            parse_prefect_file(filename, content)
        assert captured.value.code == "cell_too_long"


def test_column_profiles_contain_no_row_values() -> None:
    parsed = parse_prefect_file("prefects.csv", CSV_CONTENT)
    serialized = json.dumps([profile.__dict__ for profile in profile_columns(parsed)], ensure_ascii=False)

    assert "測試甲" not in serialized
    assert "測試乙" not in serialized
    assert "CJK_TEXT" in serialized


def test_deepseek_is_opt_in_and_reads_configuration_at_call_time(monkeypatch) -> None:
    monkeypatch.delenv("SING_YIN_DEEPSEEK_API_KEY", raising=False)
    monkeypatch.setenv("SING_YIN_DEEPSEEK_ENABLED", "false")

    status = import_assistant_status()

    assert status.enabled is False
    assert status.configured is False
    parsed = parse_prefect_file("prefects.csv", CSV_CONTENT)
    with pytest.raises(ImportAssistantError, match="disabled"):
        suggest_deepseek_column_mapping(parsed)


def test_deepseek_receives_only_headings_and_anonymous_value_kinds(monkeypatch) -> None:
    monkeypatch.setenv("SING_YIN_DEEPSEEK_ENABLED", "true")
    monkeypatch.setenv("SING_YIN_DEEPSEEK_API_KEY", "fake-local-test-key")
    monkeypatch.setenv("SING_YIN_DEEPSEEK_MODEL", "deepseek-v4-flash")
    parsed = parse_prefect_file(
        "unfamiliar.csv",
        "學生中文名,就讀級別,所屬班別,團隊身份,可服務星期\n測試甲,F.3,3A,導學風紀,星期一\n".encode("utf-8"),
    )
    captured: dict[str, object] = {}

    def transport(payload, api_key, timeout):
        captured.update(payload=payload, api_key=api_key, timeout=timeout)
        return {
            "choices": [
                {
                    "message": {
                        "content": json.dumps(
                            {
                                "mapping": [
                                    {"source": "學生中文名", "target": "name_zh"},
                                    {"source": "就讀級別", "target": "form"},
                                    {"source": "所屬班別", "target": "class_name"},
                                    {"source": "團隊身份", "target": "role"},
                                    {"source": "可服務星期", "target": "available_days"},
                                ]
                            },
                            ensure_ascii=False,
                        )
                    }
                }
            ]
        }

    suggestion = suggest_deepseek_column_mapping(parsed, transport=transport)
    request_text = json.dumps(captured["payload"], ensure_ascii=False)

    assert suggestion.suggested_target_count == 5
    assert suggestion.target_to_source["name_zh"] == "學生中文名"
    assert "測試甲" not in request_text
    assert "fake-local-test-key" not in request_text
    assert "CJK_TEXT" in request_text
    assert captured["timeout"] == 8.0


def test_deepseek_cannot_escape_the_approved_mapping_schema(monkeypatch) -> None:
    monkeypatch.setenv("SING_YIN_DEEPSEEK_ENABLED", "true")
    monkeypatch.setenv("SING_YIN_DEEPSEEK_API_KEY", "fake-local-test-key")
    parsed = parse_prefect_file("prefects.csv", CSV_CONTENT)

    def transport(_payload, _api_key, _timeout):
        return {
            "choices": [
                {"message": {"content": '{"mapping":[{"source":"姓名","target":"history_weight"}]}'}}
            ]
        }

    with pytest.raises(ImportAssistantError, match="outside the approved"):
        suggest_deepseek_column_mapping(parsed, existing_mapping={}, transport=transport)
